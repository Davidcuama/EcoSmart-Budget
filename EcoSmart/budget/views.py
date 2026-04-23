# budget/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.db.models import Sum, Avg, Count
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from datetime import datetime
from decimal import Decimal, InvalidOperation
from .models import Ingreso, Gasto, Presupuesto, Categoria, ObjetivoAhorro


# ─── Autenticación ────────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('home')
    else:
        form = AuthenticationForm()
    return render(request, 'budget/login.html', {'form': form})


def register_view(request):
    if request.user.is_authenticated:
        return redirect('home')
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('home')
    else:
        form = UserCreationForm()
    return render(request, 'budget/register.html', {'form': form})


def logout_view(request):
    logout(request)
    return redirect('login')


# ─── Home ──────────────────────────────────────────────────────────────────────

@login_required
def home(request):
    total_ingresos = Ingreso.objects.filter(usuario=request.user).aggregate(Sum('monto'))['monto__sum'] or 0
    total_gastos = Gasto.objects.filter(usuario=request.user).aggregate(Sum('monto'))['monto__sum'] or 0
    balance = float(total_ingresos) - float(total_gastos)

    return render(request, 'budget/home.html', {
        'total_ingresos': total_ingresos,
        'total_gastos': total_gastos,
        'balance': balance,
    })


# ─── Presupuesto ───────────────────────────────────────────────────────────────

@login_required
def remaining_budget(request):
    current_month = datetime.now().month
    current_year = datetime.now().year

    presupuestos = Presupuesto.objects.filter(mes=current_month, anio=current_year, usuario=request.user)

    presupuestos_con_restante = []
    for presupuesto in presupuestos:
        total_gastos = Gasto.objects.filter(
            categoria=presupuesto.categoria,
            fecha__month=current_month,
            fecha__year=current_year,
            usuario=request.user,
        ).aggregate(Sum('monto'))['monto__sum'] or 0

        restante = float(presupuesto.monto_limite) - float(total_gastos)
        porcentaje = (float(total_gastos) / float(presupuesto.monto_limite) * 100) if presupuesto.monto_limite > 0 else 0

        if porcentaje >= 100:
            estado = 'Tope Máximo'
        elif porcentaje >= 80:
            estado = 'Crítico'
        else:
            estado = 'OK'

        presupuestos_con_restante.append({
            'presupuesto': presupuesto,
            'total_gastos': total_gastos,
            'restante': restante,
            'porcentaje': round(porcentaje, 2),
            'estado': estado,
        })

    return render(request, 'budget/remaining_budget.html', {
        'presupuestos': presupuestos_con_restante,
        'mes': current_month,
        'anio': current_year,
    })


# ─── Categorías ────────────────────────────────────────────────────────────────

@login_required
def manage_categories(request):
    categorias = Categoria.objects.filter(usuario=request.user)
    return render(request, 'budget/manage_categories.html', {'categorias': categorias})


@login_required
def category_create(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        if nombre:
            Categoria.objects.create(nombre=nombre, usuario=request.user)
            return redirect('manage_categories')
    return render(request, 'budget/category_form.html', {'titulo': 'Crear Categoría', 'action': 'crear'})


@login_required
def category_edit(request, id):
    categoria = get_object_or_404(Categoria, id=id, usuario=request.user)
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        if nombre:
            categoria.nombre = nombre
            categoria.save()
            return redirect('manage_categories')
    return render(request, 'budget/category_form.html', {
        'titulo': 'Editar Categoría',
        'action': 'editar',
        'categoria': categoria,
    })


@login_required
def category_delete(request, id):
    categoria = get_object_or_404(Categoria, id=id, usuario=request.user)
    if request.method == 'POST':
        categoria.delete()
        return redirect('manage_categories')
    return render(request, 'budget/category_confirm_delete.html', {'categoria': categoria})


# ─── Ingresos ──────────────────────────────────────────────────────────────────

@login_required
def income_register(request):
    import json

    if request.method == 'POST':
        descripcion  = request.POST.get('descripcion')
        monto        = request.POST.get('monto')
        categoria_id = request.POST.get('categoria')
        categoria = Categoria.objects.filter(id=categoria_id, usuario=request.user).first() if categoria_id else None
        Ingreso.objects.create(descripcion=descripcion, monto=monto, categoria=categoria, usuario=request.user)
        return redirect('income_register')

    filtro_categoria = request.GET.get('categoria', '').strip()
    ingresos = Ingreso.objects.filter(usuario=request.user)
    if filtro_categoria:
        ingresos = ingresos.filter(categoria__id=filtro_categoria)

    categorias = Categoria.objects.filter(usuario=request.user)

    ingresos_query = Ingreso.objects.filter(usuario=request.user)
    if filtro_categoria:
        ingresos_query = ingresos_query.filter(categoria__id=filtro_categoria)

    ingresos_por_categoria = ingresos_query.values('categoria__nombre').annotate(total=Sum('monto')).order_by('-total')

    categorias_list, montos_list = [], []
    for ingreso in ingresos_por_categoria:
        categorias_list.append(ingreso['categoria__nombre'] or 'Sin categoría')
        montos_list.append(float(ingreso['total']) if ingreso['total'] else 0)

    return render(request, 'budget/income_register.html', {
        'categorias':       categorias,
        'ingresos':         ingresos,
        'filtro_categoria': filtro_categoria,
        'categorias_json':  json.dumps(categorias_list),
        'montos_json':      json.dumps(montos_list),
    })


@login_required
def income_delete(request, id):
    ingreso = get_object_or_404(Ingreso, id=id, usuario=request.user)
    if request.method == 'POST':
        ingreso.delete()
    return redirect('income_register')


# ─── Gastos ────────────────────────────────────────────────────────────────────

@login_required
def expense_record(request):
    import json

    if request.method == 'POST':
        descripcion  = request.POST.get('descripcion')
        monto        = request.POST.get('monto')
        categoria_id = request.POST.get('categoria')
        categoria = Categoria.objects.filter(id=categoria_id, usuario=request.user).first() if categoria_id else None
        Gasto.objects.create(descripcion=descripcion, monto=monto, categoria=categoria, usuario=request.user)

        # ─── Trigger Budget Alerts ─────────────────────────────────────────────
        if categoria:
            now = datetime.now()
            presupuesto = Presupuesto.objects.filter(
                categoria=categoria,
                mes=now.month,
                anio=now.year,
                usuario=request.user,
            ).first()
            if presupuesto and presupuesto.monto_limite > 0:
                total_gastado = Gasto.objects.filter(
                    categoria=categoria,
                    fecha__month=now.month,
                    fecha__year=now.year,
                    usuario=request.user,
                ).aggregate(Sum('monto'))['monto__sum'] or 0
                porcentaje = float(total_gastado) / float(presupuesto.monto_limite) * 100
                if porcentaje >= 100:
                    messages.error(
                        request,
                        f'🔴 ¡Has superado el presupuesto de "{categoria.nombre}"! '
                        f'Gastado: ${float(total_gastado):.2f} / Límite: ${float(presupuesto.monto_limite):.2f}'
                    )
                elif porcentaje >= 80:
                    messages.warning(
                        request,
                        f'⚠️ Llevas el {porcentaje:.0f}% del presupuesto de "{categoria.nombre}". '
                        f'Gastado: ${float(total_gastado):.2f} / Límite: ${float(presupuesto.monto_limite):.2f}'
                    )
        # ───────────────────────────────────────────────────────────────────────

        return redirect('expense_record')

    filtro_categoria = request.GET.get('categoria', '').strip()
    gastos = Gasto.objects.filter(usuario=request.user)
    if filtro_categoria:
        gastos = gastos.filter(categoria__id=filtro_categoria)

    categorias = Categoria.objects.filter(usuario=request.user)

    gastos_query = Gasto.objects.filter(usuario=request.user)
    if filtro_categoria:
        gastos_query = gastos_query.filter(categoria__id=filtro_categoria)

    gastos_por_categoria = gastos_query.values('categoria__nombre').annotate(total=Sum('monto')).order_by('-total')

    categorias_list, montos_list = [], []
    for gasto in gastos_por_categoria:
        categorias_list.append(gasto['categoria__nombre'] or 'Sin categoría')
        montos_list.append(float(gasto['total']) if gasto['total'] else 0)

    return render(request, 'budget/expense_record.html', {
        'categorias':       categorias,
        'gastos':           gastos,
        'filtro_categoria': filtro_categoria,
        'categorias_json':  json.dumps(categorias_list),
        'montos_json':      json.dumps(montos_list),
    })


@login_required
def expense_delete(request, id):
    gasto = get_object_or_404(Gasto, id=id, usuario=request.user)
    if request.method == 'POST':
        gasto.delete()
    return redirect('expense_record')


# ─── Historial ────────────────────────────────────────────────────────────────

@login_required
def transaction_history(request):
    filtro_tipo      = request.GET.get('tipo', '')
    filtro_categoria = request.GET.get('categoria', '').strip()
    filtro_desde     = request.GET.get('desde', '')
    filtro_hasta     = request.GET.get('hasta', '')

    ingresos = Ingreso.objects.filter(usuario=request.user)
    gastos   = Gasto.objects.filter(usuario=request.user)

    if filtro_categoria:
        ingresos = ingresos.filter(categoria__id=filtro_categoria)
        gastos   = gastos.filter(categoria__id=filtro_categoria)

    if filtro_desde:
        ingresos = ingresos.filter(fecha__gte=filtro_desde)
        gastos   = gastos.filter(fecha__gte=filtro_desde)

    if filtro_hasta:
        ingresos = ingresos.filter(fecha__lte=filtro_hasta)
        gastos   = gastos.filter(fecha__lte=filtro_hasta)

    transacciones = []

    if filtro_tipo != 'gasto':
        for i in ingresos:
            transacciones.append({
                'tipo': 'ingreso', 'descripcion': i.descripcion,
                'monto': i.monto, 'fecha': i.fecha, 'categoria': i.categoria, 'id': i.id,
            })

    if filtro_tipo != 'ingreso':
        for g in gastos:
            transacciones.append({
                'tipo': 'gasto', 'descripcion': g.descripcion,
                'monto': g.monto, 'fecha': g.fecha, 'categoria': g.categoria, 'id': g.id,
            })

    transacciones.sort(key=lambda t: t['fecha'], reverse=True)
    categorias = Categoria.objects.filter(usuario=request.user)

    return render(request, 'budget/transaction_history.html', {
        'transacciones':    transacciones,
        'categorias':       categorias,
        'filtro_tipo':      filtro_tipo,
        'filtro_categoria': int(filtro_categoria) if filtro_categoria else '',
        'filtro_desde':     filtro_desde,
        'filtro_hasta':     filtro_hasta,
    })


# ─── Estadísticas ─────────────────────────────────────────────────────────────

@login_required
def financial_statistics(request):
    from django.db.models.functions import TruncMonth

    total_ingresos = Ingreso.objects.filter(usuario=request.user).aggregate(total=Sum('monto'))['total'] or 0
    total_gastos   = Gasto.objects.filter(usuario=request.user).aggregate(total=Sum('monto'))['total'] or 0
    balance        = float(total_ingresos) - float(total_gastos)
    tasa_ahorro    = round((balance / float(total_ingresos) * 100), 2) if total_ingresos > 0 else 0

    num_ingresos = Ingreso.objects.filter(usuario=request.user).count()
    num_gastos   = Gasto.objects.filter(usuario=request.user).count()

    promedio_ingreso = round(float(total_ingresos) / num_ingresos, 2) if num_ingresos > 0 else 0
    promedio_gasto   = round(float(total_gastos) / num_gastos, 2)     if num_gastos   > 0 else 0

    cat_gastos = (
        Gasto.objects.filter(usuario=request.user)
        .values('categoria__nombre')
        .annotate(total=Sum('monto'))
        .order_by('-total')
    )
    top_categoria = cat_gastos.first() if cat_gastos.exists() else None

    ingresos_por_mes = (
        Ingreso.objects.filter(usuario=request.user)
        .annotate(mes=TruncMonth('fecha'))
        .values('mes')
        .annotate(total=Sum('monto'))
        .order_by('mes')
    )

    gastos_por_mes = (
        Gasto.objects.filter(usuario=request.user)
        .annotate(mes=TruncMonth('fecha'))
        .values('mes')
        .annotate(total=Sum('monto'))
        .order_by('mes')
    )

    meses = sorted(set(
        [r['mes'] for r in ingresos_por_mes] + [r['mes'] for r in gastos_por_mes]
    ))

    ing_dict   = {r['mes']: float(r['total']) for r in ingresos_por_mes}
    gasto_dict = {r['mes']: float(r['total']) for r in gastos_por_mes}

    evolucion = []
    for mes in meses:
        ing = ing_dict.get(mes, 0)
        gas = gasto_dict.get(mes, 0)
        evolucion.append({
            'mes': mes.strftime('%b %Y'), 'ingreso': ing,
            'gasto': gas, 'balance': round(ing - gas, 2),
        })

    mes_mayor_gasto = max(evolucion, key=lambda x: x['gasto']) if evolucion else None

    return render(request, 'budget/financial_statistics.html', {
        'total_ingresos':   total_ingresos,
        'total_gastos':     total_gastos,
        'balance':          balance,
        'tasa_ahorro':      tasa_ahorro,
        'num_ingresos':     num_ingresos,
        'num_gastos':       num_gastos,
        'promedio_ingreso': promedio_ingreso,
        'promedio_gasto':   promedio_gasto,
        'top_categoria':    top_categoria,
        'evolucion':        evolucion,
        'mes_mayor_gasto':  mes_mayor_gasto,
        'cat_gastos':       cat_gastos,
    })


# ─── Analizar hábitos de gasto ────────────────────────────────────────────────

@login_required
def analizar_habitos(request):
    import json
    from django.db.models.functions import TruncMonth

    # Gastos por categoría (todos los tiempos)
    gastos_por_cat = (
        Gasto.objects.filter(usuario=request.user)
        .values('categoria__nombre')
        .annotate(total=Sum('monto'), cantidad=Count('id'))
        .order_by('-total')
    )

    cat_labels  = [g['categoria__nombre'] or 'Sin categoría' for g in gastos_por_cat]
    cat_totales = [float(g['total']) for g in gastos_por_cat]

    # Evolución mensual de gastos (últimos 6 meses)
    gastos_mes = (
        Gasto.objects.filter(usuario=request.user)
        .annotate(mes=TruncMonth('fecha'))
        .values('mes')
        .annotate(total=Sum('monto'))
        .order_by('mes')
    )

    meses_labels  = [r['mes'].strftime('%b %Y') for r in gastos_mes]
    meses_totales = [float(r['total']) for r in gastos_mes]

    # Tendencia: comparar último mes vs mes anterior
    tendencia = None
    if len(meses_totales) >= 2:
        diff = meses_totales[-1] - meses_totales[-2]
        tendencia = {
            'diferencia': round(abs(diff), 2),
            'sube': diff > 0,
            'mes_actual': meses_labels[-1] if meses_labels else '',
            'mes_anterior': meses_labels[-2] if len(meses_labels) >= 2 else '',
        }

    # Promedio mensual de gasto
    promedio_mensual = round(sum(meses_totales) / len(meses_totales), 2) if meses_totales else 0

    # Top 3 categorías con más gasto
    top3 = list(gastos_por_cat[:3])

    # Total general
    total_gastos = sum(cat_totales)

    # Porcentaje por categoría
    for g in gastos_por_cat:
        g['porcentaje'] = round(float(g['total']) / total_gastos * 100, 1) if total_gastos > 0 else 0

    return render(request, 'budget/analizar_habitos.html', {
        'cat_labels_json':   json.dumps(cat_labels),
        'cat_totales_json':  json.dumps(cat_totales),
        'meses_labels_json': json.dumps(meses_labels),
        'meses_totales_json': json.dumps(meses_totales),
        'tendencia':         tendencia,
        'promedio_mensual':  promedio_mensual,
        'top3':              top3,
        'gastos_por_cat':    gastos_por_cat,
        'total_gastos':      total_gastos,
    })


# ─── Objetivo de ahorro ───────────────────────────────────────────────────────

@login_required
def savings_goal(request):
    error = None

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add_progress':
            objetivo_id = request.POST.get('objetivo_id')
            abono_raw   = request.POST.get('abono', '').strip()

            if not objetivo_id or not abono_raw:
                error = 'Debes indicar un objetivo y un monto para registrar avance.'
            else:
                try:
                    abono = Decimal(abono_raw)
                    if abono <= 0:
                        error = 'El abono debe ser mayor que 0.'
                    else:
                        objetivo = get_object_or_404(ObjetivoAhorro, id=objetivo_id, usuario=request.user)
                        objetivo.monto_ahorrado += abono
                        objetivo.save(update_fields=['monto_ahorrado'])
                        return redirect('savings_goal')
                except InvalidOperation:
                    error = 'El monto del abono no es válido.'
        else:
            nombre             = request.POST.get('nombre', '').strip()
            monto_objetivo_raw = request.POST.get('monto_objetivo', '').strip()
            fecha_objetivo     = request.POST.get('fecha_objetivo') or None

            if not nombre or not monto_objetivo_raw:
                error = 'Debes completar nombre y monto objetivo.'
            else:
                try:
                    monto_objetivo = Decimal(monto_objetivo_raw)
                    if monto_objetivo <= 0:
                        error = 'El monto objetivo debe ser mayor que 0.'
                    else:
                        ObjetivoAhorro.objects.create(
                            nombre=nombre,
                            monto_objetivo=monto_objetivo,
                            fecha_objetivo=fecha_objetivo,
                            usuario=request.user,
                        )
                        return redirect('savings_goal')
                except InvalidOperation:
                    error = 'El monto objetivo no es válido.'

    objetivos_db = ObjetivoAhorro.objects.filter(usuario=request.user).order_by('-fecha_creacion', '-id')
    objetivos = []
    for objetivo in objetivos_db:
        monto_objetivo = float(objetivo.monto_objetivo)
        monto_ahorrado = float(objetivo.monto_ahorrado)
        progreso       = round((monto_ahorrado / monto_objetivo) * 100, 2) if monto_objetivo > 0 else 0
        objetivos.append({
            'obj':            objetivo,
            'progreso':       progreso,
            'progreso_visual': min(progreso, 100),
            'restante':       round(monto_objetivo - monto_ahorrado, 2),
            'completado':     monto_ahorrado >= monto_objetivo,
        })

    return render(request, 'budget/savings_goal.html', {'objetivos': objetivos, 'error': error})


# ─── Crear presupuesto ────────────────────────────────────────────────────────

@login_required
def budget_create(request):
    if request.method == 'POST':
        categoria_id = request.POST.get('categoria')
        monto_limite = request.POST.get('monto_limite')
        mes          = request.POST.get('mes')
        anio         = request.POST.get('anio')
        categoria = get_object_or_404(Categoria, id=categoria_id, usuario=request.user)
        Presupuesto.objects.create(
            categoria=categoria, monto_limite=monto_limite,
            mes=mes, anio=anio, usuario=request.user,
        )
        return redirect('budget_create')

    categorias   = Categoria.objects.filter(usuario=request.user)
    presupuestos = Presupuesto.objects.filter(usuario=request.user)
    return render(request, 'budget/budget_create.html', {
        'categorias': categorias, 'presupuestos': presupuestos,
    })


# ─── Exportar PDF / Excel ─────────────────────────────────────────────────────

@login_required
def export_pdfpage(request):
    current_month = datetime.now().month
    current_year  = datetime.now().year
    years = list(range(current_year - 5, current_year + 3))
    return render(request, 'budget/export_pdf.html', {
        'current_month': current_month,
        'current_year':  current_year,
        'years':         years,
    })


@login_required
def export_monthly_pdf(request):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.units import inch

    mes  = request.GET.get('mes', datetime.now().month)
    anio = request.GET.get('anio', datetime.now().year)
    try:
        mes  = int(mes)
        anio = int(anio)
    except (ValueError, TypeError):
        mes  = datetime.now().month
        anio = datetime.now().year

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_{mes:02d}_{anio}.pdf"'

    doc    = SimpleDocTemplate(response, pagesize=letter)
    story  = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'],
        fontSize=24, textColor=colors.HexColor('#1a472a'),
        spaceAfter=30, alignment=1,
    )
    heading_style = ParagraphStyle(
        'CustomHeading', parent=styles['Heading2'],
        fontSize=14, textColor=colors.HexColor('#2d7a4a'),
        spaceAfter=12, spaceBefore=12,
    )

    mes_nombre = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                  'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][mes - 1]
    story.append(Paragraph(f'Reporte Financiero - {mes_nombre} {anio}', title_style))
    story.append(Spacer(1, 0.3 * inch))

    # Presupuestos
    story.append(Paragraph('Presupuestos vs Gastos', heading_style))
    presupuestos = Presupuesto.objects.filter(mes=mes, anio=anio, usuario=request.user)

    if presupuestos.exists():
        datos = [['Categoría', 'Límite', 'Gasto', 'Disponible', 'Estado']]
        for p in presupuestos:
            gasto_total = Gasto.objects.filter(
                categoria=p.categoria, fecha__month=mes, fecha__year=anio, usuario=request.user,
            ).aggregate(Sum('monto'))['monto__sum'] or 0
            disponible = float(p.monto_limite) - float(gasto_total)
            datos.append([str(p.categoria), f"${p.monto_limite:,.2f}", f"${gasto_total:,.2f}",
                          f"${disponible:,.2f}", 'EXCEDIDO' if disponible < 0 else 'OK'])
        t = Table(datos, colWidths=[1.8 * inch, 1.1 * inch, 1.1 * inch, 1.1 * inch, 0.9 * inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d7a4a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        story.append(t)
    else:
        story.append(Paragraph('No hay presupuestos registrados para este mes.', styles['Normal']))

    story.append(Spacer(1, 0.2 * inch))

    # Ingresos
    story.append(Paragraph('Ingresos del Mes', heading_style))
    ingresos       = Ingreso.objects.filter(fecha__month=mes, fecha__year=anio, usuario=request.user)
    total_ingresos = ingresos.aggregate(Sum('monto'))['monto__sum'] or 0

    if ingresos.exists():
        datos = [['Descripción', 'Categoría', 'Monto', 'Fecha']]
        for i in ingresos:
            datos.append([i.descripcion, str(i.categoria) if i.categoria else 'Sin categoría',
                          f"${i.monto:,.2f}", i.fecha.strftime('%d/%m/%Y')])
        datos.append(['TOTAL', '', f"${total_ingresos:,.2f}", ''])
        t = Table(datos, colWidths=[2.2 * inch, 1.5 * inch, 1.3 * inch, 1.2 * inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d7a4a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#e8f5e9')),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#c8e6c9')),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        story.append(t)
    else:
        story.append(Paragraph('No hay ingresos registrados para este mes.', styles['Normal']))

    story.append(Spacer(1, 0.2 * inch))

    # Gastos
    story.append(Paragraph('Gastos del Mes', heading_style))
    gastos       = Gasto.objects.filter(fecha__month=mes, fecha__year=anio, usuario=request.user)
    total_gastos = gastos.aggregate(Sum('monto'))['monto__sum'] or 0

    if gastos.exists():
        datos = [['Descripción', 'Categoría', 'Monto', 'Fecha']]
        for g in gastos:
            datos.append([g.descripcion, str(g.categoria) if g.categoria else 'Sin categoría',
                          f"${g.monto:,.2f}", g.fecha.strftime('%d/%m/%Y')])
        datos.append(['TOTAL', '', f"${total_gastos:,.2f}", ''])
        t = Table(datos, colWidths=[2.2 * inch, 1.5 * inch, 1.3 * inch, 1.2 * inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c62828')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#ffebee')),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ef9a9a')),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        story.append(t)
    else:
        story.append(Paragraph('No hay gastos registrados para este mes.', styles['Normal']))

    story.append(Spacer(1, 0.2 * inch))

    # Resumen
    story.append(Paragraph('Resumen Mensual', heading_style))
    balance = float(total_ingresos) - float(total_gastos)
    t = Table([
        ['Total Ingresos', f"${total_ingresos:,.2f}"],
        ['Total Gastos',   f"${total_gastos:,.2f}"],
        ['Balance',        f"${balance:,.2f}"],
    ], colWidths=[3 * inch, 2 * inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f5f5f5')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
    ]))
    story.append(t)
    story.append(PageBreak())

    # Objetivos de ahorro
    story.append(Paragraph('Objetivos de Ahorro y Progreso', heading_style))
    objetivos = ObjetivoAhorro.objects.filter(usuario=request.user)

    if objetivos.exists():
        datos = [['Objetivo', 'Meta', 'Ahorrado', 'Progreso', 'Restante']]
        for obj in objetivos:
            mo = float(obj.monto_objetivo)
            ma = float(obj.monto_ahorrado)
            progreso = round(ma / mo * 100, 2) if mo > 0 else 0
            datos.append([obj.nombre, f"${mo:,.2f}", f"${ma:,.2f}", f"{progreso}%",
                          f"${round(mo - ma, 2):,.2f}"])
        t = Table(datos, colWidths=[1.8 * inch, 1.2 * inch, 1.2 * inch, 1.2 * inch, 1.2 * inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d7a4a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f0fdf0')),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        story.append(t)
    else:
        story.append(Paragraph('No hay objetivos de ahorro registrados.', styles['Normal']))

    doc.build(story)
    return response


@login_required
def export_monthly_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    mes  = request.GET.get('mes', datetime.now().month)
    anio = request.GET.get('anio', datetime.now().year)
    try:
        mes  = int(mes)
        anio = int(anio)
    except (ValueError, TypeError):
        mes  = datetime.now().month
        anio = datetime.now().year

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte'

    titulo_font  = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    titulo_fill  = PatternFill(start_color='1a472a', end_color='1a472a', fill_type='solid')
    header_font  = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill  = PatternFill(start_color='2d7a4a', end_color='2d7a4a', fill_type='solid')
    total_font   = Font(name='Calibri', size=11, bold=True)
    total_fill   = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    border       = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    mes_nombre = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                  'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][mes - 1]

    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value     = f'Reporte Financiero - {mes_nombre} {anio}'
    cell.font      = titulo_font
    cell.fill      = titulo_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    row = 3

    def section_header(label, fill=header_fill):
        nonlocal row
        ws.merge_cells(f'A{row}:E{row}')
        c = ws[f'A{row}']
        c.value     = label
        c.font      = header_font
        c.fill      = fill
        c.alignment = Alignment(horizontal='left', vertical='center')
        row += 1

    def add_row(values, fmt=None):
        nonlocal row
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col)
            c.value     = val
            c.border    = border
            c.alignment = Alignment(horizontal='center')
            if fmt and fmt.get(col):
                c.number_format = fmt[col]
        row += 1

    # Presupuestos
    section_header('Presupuestos vs Gastos')
    presupuestos = Presupuesto.objects.filter(mes=mes, anio=anio, usuario=request.user)
    if presupuestos.exists():
        add_row(['Categoría', 'Límite', 'Gasto', 'Disponible', 'Estado'])
        for p in presupuestos:
            gt = Gasto.objects.filter(
                categoria=p.categoria, fecha__month=mes, fecha__year=anio, usuario=request.user,
            ).aggregate(Sum('monto'))['monto__sum'] or 0
            disp = float(p.monto_limite) - float(gt)
            add_row([str(p.categoria), float(p.monto_limite), float(gt), disp,
                     'EXCEDIDO' if disp < 0 else 'OK'],
                    {2: '$#,##0.00', 3: '$#,##0.00', 4: '$#,##0.00'})
    else:
        ws.merge_cells(f'A{row}:E{row}')
        ws.cell(row=row, column=1).value = 'No hay presupuestos registrados para este mes.'
        row += 1
    row += 1

    # Ingresos
    section_header('Ingresos del Mes')
    ingresos       = Ingreso.objects.filter(fecha__month=mes, fecha__year=anio, usuario=request.user)
    total_ingresos = ingresos.aggregate(Sum('monto'))['monto__sum'] or 0
    if ingresos.exists():
        add_row(['Descripción', 'Categoría', 'Monto', 'Fecha'])
        for i in ingresos:
            add_row([i.descripcion, str(i.categoria) if i.categoria else 'Sin categoría',
                     float(i.monto), i.fecha], {3: '$#,##0.00', 4: 'dd/mm/yyyy'})
        add_row(['TOTAL', '', float(total_ingresos), ''], {3: '$#,##0.00'})
    else:
        ws.merge_cells(f'A{row}:E{row}')
        ws.cell(row=row, column=1).value = 'No hay ingresos registrados para este mes.'
        row += 1
    row += 1

    # Gastos
    section_header('Gastos del Mes', PatternFill(start_color='C62828', end_color='C62828', fill_type='solid'))
    gastos       = Gasto.objects.filter(fecha__month=mes, fecha__year=anio, usuario=request.user)
    total_gastos = gastos.aggregate(Sum('monto'))['monto__sum'] or 0
    if gastos.exists():
        add_row(['Descripción', 'Categoría', 'Monto', 'Fecha'])
        for g in gastos:
            add_row([g.descripcion, str(g.categoria) if g.categoria else 'Sin categoría',
                     float(g.monto), g.fecha], {3: '$#,##0.00', 4: 'dd/mm/yyyy'})
        add_row(['TOTAL', '', float(total_gastos), ''], {3: '$#,##0.00'})
    else:
        ws.merge_cells(f'A{row}:E{row}')
        ws.cell(row=row, column=1).value = 'No hay gastos registrados para este mes.'
        row += 1
    row += 1

    # Resumen
    section_header('Resumen Mensual')
    balance = float(total_ingresos) - float(total_gastos)
    for desc, val in [('Total Ingresos', float(total_ingresos)),
                      ('Total Gastos',   float(total_gastos)),
                      ('Balance',        balance)]:
        ws.cell(row=row, column=1).value = desc
        ws.cell(row=row, column=2).value = val
        for col in (1, 2):
            ws.cell(row=row, column=col).font   = total_font
            ws.cell(row=row, column=col).fill   = total_fill
            ws.cell(row=row, column=col).border = border
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        row += 1
    row += 2

    # Objetivos
    section_header('Objetivos de Ahorro y Progreso')
    objetivos = ObjetivoAhorro.objects.filter(usuario=request.user)
    if objetivos.exists():
        add_row(['Objetivo', 'Meta', 'Ahorrado', 'Progreso %', 'Restante'])
        for obj in objetivos:
            mo = float(obj.monto_objetivo)
            ma = float(obj.monto_ahorrado)
            pr = round(ma / mo * 100, 2) if mo > 0 else 0
            add_row([obj.nombre, mo, ma, pr, round(mo - ma, 2)],
                    {2: '$#,##0.00', 3: '$#,##0.00', 4: '0.00', 5: '$#,##0.00'})
    else:
        ws.merge_cells(f'A{row}:E{row}')
        ws.cell(row=row, column=1).value = 'No hay objetivos de ahorro registrados.'

    for col, w in zip('ABCDE', [25, 15, 15, 15, 15]):
        ws.column_dimensions[col].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_{mes:02d}_{anio}.xlsx"'
    wb.save(response)
    return response
